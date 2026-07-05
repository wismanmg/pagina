# -*- coding: utf-8 -*-
"""Generación de reportes Excel (por enlace y de avance)."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL = PatternFill("solid", start_color="1F4E79")
AZUL2 = PatternFill("solid", start_color="2E75B6")
GRIS = PatternFill("solid", start_color="F2F2F2")
VERDE = PatternFill("solid", start_color="70AD47")
ROJO = PatternFill("solid", start_color="F8CBAD")
_thin = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(ws, fila, textos, fill=AZUL, col0=1):
    for i, t in enumerate(textos):
        c = ws.cell(fila, col0 + i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE


def excel_enlace(e):
    wb = Workbook(); ws = wb.active; ws.title = "REPORTE"
    ws["A1"] = f"{e.anillo} | {e.nombre}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = (f"Capacidad: {e.capacidad} · Tipo: {e.tipo_cable or '-'} · "
                f"Longitud: {e.longitud or '-'} · Estado: {e.estado} · "
                f"Creado: {e.creado:%d/%m/%Y %H:%M}")
    ws["A3"] = (f"EXTREMO A: {e.origen_a or '?'} | Sala: {e.sala_a or '-'} | "
                f"Rack: {e.rack_a or '-'} | Posición: {e.posicion_a or '-'}")
    ws["A4"] = (f"EXTREMO B: {e.origen_b or '?'} | Sala: {e.sala_b or '-'} | "
                f"Rack: {e.rack_b or '-'} | Posición: {e.posicion_b or '-'}")
    for cel in ("A2", "A3", "A4"):
        ws[cel].font = Font(size=9, color="595959")

    ws.append([])
    _hdr(ws, 6, ["FIBRA", f"DESCRIPCIÓN {e.origen_a or 'A'}",
                 f"DESCRIPCIÓN {e.origen_b or 'B'}", "COINCIDE"])
    for h in e.hilos:
        ws.append([h.numero, h.descripcion_a, h.descripcion_b,
                   "Sí" if h.coincide else "No"])
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(r, c).border = BORDE
        if not h.coincide:
            ws.cell(r, 4).fill = ROJO
    for col, w in zip("ABCD", [8, 50, 50, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, f"{e.nombre.replace(' ', '_')}.xlsx"


def excel_avance(enlaces):
    ahora = datetime.utcnow()
    pend = [e for e in enlaces if e.estado == "INCOMPLETO"]
    valid = [e for e in enlaces if e.estado == "VALIDADO"]
    total = len(enlaces)

    wb = Workbook(); rs = wb.active; rs.title = "RESUMEN"
    rs["A1"] = "REPORTE DE AVANCE — VALIDACIÓN DE ENLACES"
    rs["A1"].font = Font(bold=True, size=14, color="1F4E79")
    rs["A2"] = f"Generado: {ahora:%d/%m/%Y %H:%M} (UTC)"
    rs["A2"].font = Font(size=9, color="595959")

    kpis = [
        ("Total de enlaces", total),
        ("Enlaces validados (A y B)", len(valid)),
        ("Enlaces pendientes", len(pend)),
        ("% de avance", f"{(len(valid)/total*100):.1f}%" if total else "0%"),
        ("Hilos validados", sum(e.capacidad for e in valid)),
        ("Hilos pendientes", sum(e.capacidad for e in pend)),
        ("Pendiente más antiguo (días)", max((e.dias for e in pend), default=0)),
        ("Antigüedad promedio (días)",
         round(sum(e.dias for e in pend)/len(pend)) if pend else 0),
        ("Pendientes solo por Extremo B",
         sum(1 for e in pend if e.a_completo and not e.b_completo)),
        ("Pendientes solo por Extremo A",
         sum(1 for e in pend if e.b_completo and not e.a_completo)),
        ("Pendientes por ambos extremos",
         sum(1 for e in pend if not e.a_completo and not e.b_completo)),
        ("Pendientes en semáforo ROJO (≥90 días)",
         sum(1 for e in pend if e.semaforo == "rojo")),
    ]
    _hdr(rs, 4, ["INDICADOR", "VALOR"], AZUL2)
    for i, (k, v) in enumerate(kpis):
        r = 5 + i
        rs.cell(r, 1, k).border = BORDE
        c = rs.cell(r, 2, v); c.border = BORDE
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")

    _hdr(rs, 4, ["ANILLO", "PEND.", "VALID.", "% AVANCE"], AZUL2, col0=4)
    for i, an in enumerate(["ANILLO NORTE", "ANILLO SUR",
                            "ANILLO ESTE", "ANILLO OESTE"]):
        r = 5 + i
        np_ = sum(1 for e in pend if e.anillo == an)
        nv = sum(1 for e in valid if e.anillo == an)
        pct = f"{(nv/(np_+nv)*100):.0f}%" if (np_ + nv) else "-"
        for j, v in enumerate([an, np_, nv, pct]):
            c = rs.cell(r, 4 + j, v); c.border = BORDE
            if j: c.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFG", [38, 12, 3, 18, 9, 9, 11]):
        rs.column_dimensions[col].width = w

    wp = wb.create_sheet("PENDIENTES")
    _hdr(wp, 1, ["N°", "ANILLO", "ENLACE", "TRAMO", "CAP.", "CREADO",
                 "DÍAS", "SEMÁFORO", "QUÉ FALTA", "MODIFICADO POR"])
    for i, e in enumerate(sorted(pend, key=lambda x: x.creado), 1):
        wp.append([i, e.anillo, e.nombre,
                   f"{e.origen_a or '?'} → {e.origen_b or '?'}", e.capacidad,
                   e.creado.strftime("%d/%m/%Y %H:%M"), e.dias,
                   e.semaforo.upper(), e.pendiente_texto, e.modificado_por or "-"])
        for c in range(1, 11):
            cell = wp.cell(i + 1, c); cell.border = BORDE; cell.font = Font(size=9)
            if i % 2 == 0: cell.fill = GRIS
        if e.semaforo == "rojo":
            wp.cell(i + 1, 8).fill = ROJO
    for col, w in zip("ABCDEFGHIJ", [4, 14, 28, 28, 6, 15, 6, 10, 22, 14]):
        wp.column_dimensions[col].width = w
    wp.freeze_panes = "A2"
    wp.auto_filter.ref = f"A1:J{len(pend)+1}"

    wv = wb.create_sheet("VALIDADOS")
    _hdr(wv, 1, ["N°", "ANILLO", "ENLACE", "TRAMO", "CAP.", "CREADO",
                 "HILOS CON DIFERENCIAS"], VERDE)
    for i, e in enumerate(sorted(valid, key=lambda x: x.creado), 1):
        difs = sum(1 for h in e.hilos if not h.coincide)
        wv.append([i, e.anillo, e.nombre, f"{e.origen_a} ⇄ {e.origen_b}",
                   e.capacidad, e.creado.strftime("%d/%m/%Y %H:%M"), difs])
        for c in range(1, 8):
            wv.cell(i + 1, c).border = BORDE
            wv.cell(i + 1, c).font = Font(size=9)
    for col, w in zip("ABCDEFG", [4, 14, 28, 28, 6, 15, 20]):
        wv.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, f"REPORTE_AVANCE_{ahora:%Y%m%d}.xlsx"
