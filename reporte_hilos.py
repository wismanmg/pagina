# -*- coding: utf-8 -*-
"""
VALIDACIÓN DE ENLACES — Reporte de Hilos (reporte-hilos)
=========================================================
Aplicación completa en UN SOLO archivo Python (Flask + SQLite).

Reproduce el dashboard de https://reporte-hilos.onrender.com/ :
  - Tarjeta "Pendientes (Incompletos)" y tarjeta "Validados"
  - Botones: Completar B / Editar / Exportar / Ver / + Nuevo Enlace
  - Export a Excel por enlace (openpyxl)
  - Los 27 enlaces pendientes y 3 validados se precargan como datos iniciales

CÓMO EJECUTAR:
    pip install flask flask-sqlalchemy openpyxl
    python reporte_hilos.py
    → abrir http://127.0.0.1:5000
"""

import io
import uuid
from datetime import datetime

from flask import (Flask, request, redirect, url_for, send_file,
                   render_template_string)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///enlaces.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

# ============================================================
# 1) MODELOS
# ============================================================
class Enlace(db.Model):
    id = db.Column(db.String(36), primary_key=True,
                   default=lambda: str(uuid.uuid4()))
    anillo = db.Column(db.String(30), nullable=False)
    nombre = db.Column(db.String(120), nullable=False)
    capacidad = db.Column(db.Integer, nullable=False, default=24)
    origen_a = db.Column(db.String(80), default="")
    origen_b = db.Column(db.String(80), default="")
    estado = db.Column(db.String(15), default="INCOMPLETO")   # INCOMPLETO|VALIDADO
    creado = db.Column(db.DateTime, default=datetime.utcnow)
    hilos = db.relationship("Hilo", backref="enlace",
                            cascade="all, delete-orphan", order_by="Hilo.numero")


class Hilo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    enlace_id = db.Column(db.String(36), db.ForeignKey("enlace.id"))
    numero = db.Column(db.Integer)
    descripcion_a = db.Column(db.Text, default="")
    descripcion_b = db.Column(db.Text, default="")

    @property
    def coincide(self):
        na = " ".join((self.descripcion_a or "").upper().split())
        nb = " ".join((self.descripcion_b or "").upper().split())
        return na == nb


# ============================================================
# 2) DATOS INICIALES (los de tu dashboard actual)
#    (anillo, nombre, tramo_a, tramo_b, capacidad, fecha, estado)
# ============================================================
SEED = [
    ("ANILLO ESTE",  "SANTA ANITA - LA MOLINA",   "SANTA ANITA", "LA MOLINA",   48, "2026-04-21 17:18:46", "INCOMPLETO"),
    ("ANILLO SUR",   "LURIN - TELEPUERTO - 96SM", "LURIN",       "96SM",        96, "2026-04-18 20:10:17", "INCOMPLETO"),
    ("ANILLO ESTE",  "AVIACION - HIGUERETA",      "AVIACION",    "HIGUERETA",   96, "2026-03-12 02:40:42", "INCOMPLETO"),
    ("ANILLO ESTE",  "CAMACHO - MONTERICO",       "MONTERICO",   "CAMACHO",     48, "2026-03-08 00:52:18", "INCOMPLETO"),
    ("ANILLO ESTE",  "SAN LUIS - MOTERRICO",      "MONTERICO",   "SAN LUIS",    48, "2026-03-07 02:30:07", "INCOMPLETO"),
    ("ANILLO ESTE",  "HIGUERETA - SAN LUIS",      "HIGUERETA",   "SAN LUIS",    48, "2026-02-05 19:00:47", "INCOMPLETO"),
    ("ANILLO ESTE",  "HIGUERETA - TELEPUERTO",    "HIGUERETA",   "TELEPUERTO",  48, "2026-02-05 14:13:29", "INCOMPLETO"),
    ("ANILLO SUR",   "OCHARAN - HIGUERETA",       "OCHARAN",     "HIGUERETA",   96, "2026-02-04 21:05:33", "INCOMPLETO"),
    ("ANILLO OESTE", "SAN MIGUEL - SAN FELIPE 1", "POP SAN MIGUEL", "SAN FELIPE 1", 48, "2026-01-21 20:40:26", "INCOMPLETO"),
    ("ANILLO NORTE", "COTABAMBAS - ZARATE",       "COTABAMBAS",  "ZARATE",      96, "2026-01-20 22:26:54", "INCOMPLETO"),
    ("ANILLO NORTE", "COLONIAL - COTABAMBAS",     "COLONIAL",    "COTABAMBAS",  48, "2026-01-20 17:16:31", "INCOMPLETO"),
    ("ANILLO OESTE", "AEROPUERTO - CALLAO",       "AEROPUERTO",  "CALLAO",      96, "2026-01-19 20:28:32", "INCOMPLETO"),
    ("ANILLO SUR",   "AERIPUETO - LA MILLA",      "AERIPUETO",   "LA MILLA",    24, "2026-01-19 17:26:43", "INCOMPLETO"),
    ("ANILLO SUR",   "AEROPUERTO - PAMPA LIBRE",  "AEROPUERTO",  "PAMPA LIBRE", 24, "2026-01-19 14:28:30", "INCOMPLETO"),
    ("ANILLO NORTE", "AEROPUERTO - COLONIAL",     "AEROPUERTO",  "COLONIAL",    48, "2026-01-15 16:10:38", "INCOMPLETO"),
    ("ANILLO SUR",   "AEROPUERTO - POLO 1",       "AEROPUERTO",  "POLO 1",      24, "2026-01-15 02:07:20", "INCOMPLETO"),
    ("ANILLO OESTE", "LA PUNTA - SAN MIGUEL",     "LA PUNTA",    "SAN MIGUEL",  48, "2026-01-14 23:20:24", "INCOMPLETO"),
    ("ANILLO OESTE", "CALLO - LA PUNTA",          "CALLAO",      "LA PUNTA",    48, "2026-01-14 22:34:45", "INCOMPLETO"),
    ("ANILLO SUR",   "TELEPUERTO - ASIA",         "TELEPUERTO",  "ASIA",        24, "2026-01-14 02:16:00", "INCOMPLETO"),
    ("ANILLO NORTE", "SANTA LUZMILA - LOS OLIVOS","SANTA LUZMILA","LOS OLIVOS", 48, "2026-01-12 18:38:00", "INCOMPLETO"),
    ("ANILLO NORTE", "AEROPUERTO - LOS OLIVOS",   "AEROPUERTO",  "LOS OLIVOS",  48, "2026-01-12 16:40:00", "INCOMPLETO"),
    ("ANILLO ESTE",  "ZARATE - SANTA ANITA",      "ZARATE",      "SANTA ANITA", 48, "2026-01-10 19:30:00", "INCOMPLETO"),
    ("ANILLO NORTE", "ZARATE - INGENIERIA",       "ZARATE",      "INGENIERIA",  48, "2026-01-10 17:58:00", "INCOMPLETO"),
    ("ANILLO SUR",   "TELPUERTO -POLO 2",         "POLO 2",      "VILLA EL SAVADOR", 24, "2026-01-09 23:12:00", "INCOMPLETO"),
    ("ANILLO SUR",   "POLO 2 - POLO 1",           "P0L0 1",      "POLO 2",      24, "2026-01-09 18:01:00", "INCOMPLETO"),
    ("ANILLO SUR",   "CHORRILLOS - OCHARAN",      "CHORRILLOS",  "OCHARAN",     96, "2026-01-08 18:39:00", "INCOMPLETO"),
    ("ANILLO SUR",   "SAN JUAN - TELPUERTO",      "SAN JUAN",    "TELPUERTO",   96, "2026-01-07 22:31:00", "INCOMPLETO"),
    # ---- Validados ----
    ("ANILLO SUR",   "LURIN - TELEPUERTO - 48SM", "TELEPUERTO",  "LURIN",       48, "2026-01-14 03:15:26", "VALIDADO"),
    ("ANILLO SUR",   "SAN JUAN - POLO 2",         "SAN JUAN",    "POLO 2",      96, "2026-01-08 00:14:39", "VALIDADO"),
    ("ANILLO SUR",   "SAN JUAN - CHORRILLOS",     "SAN JUAN",    "CHORRILLOS",  48, "2026-01-07 20:04:48", "VALIDADO"),
]


def seed_inicial():
    """Precarga los enlaces del dashboard solo si la BD está vacía."""
    if Enlace.query.count() > 0:
        return
    for anillo, nombre, a, b, cap, fecha, estado in SEED:
        e = Enlace(anillo=anillo, nombre=nombre, capacidad=cap,
                   origen_a=a, origen_b=b, estado=estado,
                   creado=datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S"))
        for n in range(1, cap + 1):
            e.hilos.append(Hilo(numero=n,
                                descripcion_a="LIBRE",
                                descripcion_b="LIBRE" if estado == "VALIDADO" else ""))
        db.session.add(e)
    db.session.commit()


with app.app_context():
    db.create_all()
    seed_inicial()

# ============================================================
# 3) PLANTILLA DEL DASHBOARD (mismo diseño de tu captura)
# ============================================================
TPL_DASH = """
<!doctype html><html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard - Validación</title>
<style>
 body{margin:0;font-family:'Segoe UI',system-ui,sans-serif;background:#fff;color:#1f2937}
 header{display:flex;align-items:center;justify-content:space-between;
        padding:14px 24px;border-bottom:1px solid #e5e7eb}
 header h1{font-size:1.15rem;margin:0}
 .btn{display:inline-block;padding:6px 12px;border-radius:6px;border:1px solid #d1d5db;
      background:#fff;color:#374151;text-decoration:none;font-size:.8rem;cursor:pointer}
 .btn-azul{background:#2563eb;border-color:#2563eb;color:#fff}
 .btn-borde-azul{border-color:#2563eb;color:#2563eb}
 .btn-borde-verde{border-color:#059669;color:#059669}
 .grid{display:grid;grid-template-columns:1fr 1fr;gap:24px;padding:24px;
       max-width:1500px;margin:0 auto;align-items:start}
 .card{background:#fff;border:1px solid #e5e7eb;border-radius:12px;
       box-shadow:0 1px 4px rgba(0,0,0,.06);padding:18px}
 .card h2{font-size:1rem;margin:0 0 2px;display:flex;justify-content:space-between}
 .badge{background:#dbeafe;color:#1d4ed8;border-radius:10px;padding:1px 9px;font-size:.75rem}
 .sub{color:#6b7280;font-size:.75rem;margin:0 0 12px}
 table{width:100%;border-collapse:collapse;font-size:.8rem}
 th{text-align:left;padding:6px 4px;border-bottom:2px solid #e5e7eb}
 td{padding:9px 4px;border-bottom:1px solid #f3f4f6;vertical-align:middle}
 .tramo{color:#9ca3af;font-size:.72rem}
 .acc{display:flex;gap:5px;flex-wrap:wrap}
</style></head><body>
<header>
  <h1>🧵 Validación de Enlaces</h1>
  <a class="btn btn-azul" href="{{ url_for('nuevo') }}">+ Nuevo Enlace</a>
</header>
<div class="grid">

  <div class="card">
    <h2>Pendientes (Incompletos) <span class="badge">{{ pend|length }}</span></h2>
    <p class="sub">Enlaces creados en Extremo A y pendientes de completar Extremo B.</p>
    <table>
      <tr><th>Enlace</th><th>Cap.</th><th>Creado</th><th>Acciones</th></tr>
      {% for e in pend %}
      <tr>
        <td><strong>{{ e.anillo }} | {{ e.nombre }}</strong><br>
            <span class="tramo">{{ e.origen_a }} → {{ e.origen_b }}</span></td>
        <td>{{ e.capacidad }}</td>
        <td>{{ e.creado.strftime('%Y-%m-%d %H:%M:%S') }}</td>
        <td class="acc">
          <a class="btn btn-borde-azul" href="{{ url_for('completar_b', id=e.id) }}">Completar B</a>
          <a class="btn" href="{{ url_for('editar', id=e.id) }}">Editar</a>
          <a class="btn btn-borde-verde" href="{{ url_for('exportar', id=e.id) }}">Exportar</a>
        </td>
      </tr>
      {% endfor %}
    </table>
  </div>

  <div class="card">
    <h2>Validados <span class="badge">{{ valid|length }}</span></h2>
    <p class="sub">Enlaces con Extremo A y B completados (reporte comparativo).</p>
    <table>
      <tr><th>Enlace</th><th>Cap.</th><th>Creado</th><th>Acciones</th></tr>
      {% for e in valid %}
      <tr>
        <td><strong>{{ e.anillo }} | {{ e.nombre }}</strong><br>
            <span class="tramo">{{ e.origen_a }} ⇄ {{ e.origen_b }}</span></td>
        <td>{{ e.capacidad }}</td>
        <td>{{ e.creado.strftime('%Y-%m-%d %H:%M:%S') }}</td>
        <td class="acc">
          <a class="btn" href="{{ url_for('ver', id=e.id) }}">Ver</a>
          <a class="btn" href="{{ url_for('editar', id=e.id) }}">Editar</a>
          <a class="btn btn-borde-verde" href="{{ url_for('exportar', id=e.id) }}">Exportar</a>
        </td>
      </tr>
      {% endfor %}
    </table>
  </div>

</div></body></html>
"""

TPL_FORM = """
<!doctype html><html lang="es"><head><meta charset="utf-8">
<title>{{ titulo }}</title>
<style>
 body{font-family:'Segoe UI',sans-serif;max-width:900px;margin:24px auto;color:#1f2937}
 label{display:block;font-size:.8rem;color:#6b7280;margin-top:10px}
 input,select,textarea{width:100%;padding:6px;border:1px solid #d1d5db;border-radius:6px;box-sizing:border-box}
 table{width:100%;border-collapse:collapse;font-size:.8rem;margin-top:14px}
 th,td{border:1px solid #e5e7eb;padding:5px}
 th{background:#f3f4f6}
 .btn{display:inline-block;margin-top:16px;padding:8px 16px;border-radius:6px;
      background:#2563eb;color:#fff;border:0;cursor:pointer;text-decoration:none}
 .rojo{background:#fee2e2}
</style></head><body>
<h2>{{ titulo }}</h2>
{{ cuerpo|safe }}
</body></html>
"""

# ============================================================
# 4) RUTAS
# ============================================================
@app.route("/")
def dashboard():
    pend = Enlace.query.filter_by(estado="INCOMPLETO").order_by(Enlace.creado.desc()).all()
    valid = Enlace.query.filter_by(estado="VALIDADO").order_by(Enlace.creado.desc()).all()
    return render_template_string(TPL_DASH, pend=pend, valid=valid)


@app.route("/nuevo", methods=["GET", "POST"])
def nuevo():
    if request.method == "POST":
        cap = int(request.form.get("capacidad") or 24)
        e = Enlace(anillo=request.form.get("anillo", ""),
                   nombre=request.form.get("nombre", ""),
                   capacidad=cap,
                   origen_a=request.form.get("origen_a", ""),
                   origen_b=request.form.get("origen_b", ""))
        for n in range(1, cap + 1):
            e.hilos.append(Hilo(numero=n,
                                descripcion_a=request.form.get(f"h{n}", "")))
        db.session.add(e); db.session.commit()
        return redirect(url_for("dashboard"))

    filas = "".join(f"<tr><td>{n}</td><td><input name='h{n}' "
                    f"placeholder='LIBRE / ODF / SIN PIGTAIL...'></td></tr>"
                    for n in range(1, 25))
    cuerpo = f"""
    <form method="post">
      <label>Anillo</label>
      <select name="anillo">
        <option>ANILLO NORTE</option><option>ANILLO SUR</option>
        <option>ANILLO ESTE</option><option>ANILLO OESTE</option>
      </select>
      <label>Nombre del enlace</label><input name="nombre" required>
      <label>Capacidad (hilos)</label>
      <select name="capacidad"><option>24</option><option selected>48</option><option>96</option></select>
      <label>Origen A</label><input name="origen_a">
      <label>Origen B (sugerido)</label><input name="origen_b">
      <table><tr><th>#</th><th>Descripción Extremo A</th></tr>{filas}</table>
      <button class="btn">Guardar (Incompleto)</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM, titulo="Nuevo Enlace", cuerpo=cuerpo)


@app.route("/enlace/<id>/b", methods=["GET", "POST"])
def completar_b(id):
    e = Enlace.query.get_or_404(id)
    if request.method == "POST":
        e.origen_b = request.form.get("origen_b", e.origen_b)
        for h in e.hilos:
            h.descripcion_b = request.form.get(f"h{h.numero}", "")
        e.estado = "VALIDADO"
        db.session.commit()
        return redirect(url_for("ver", id=e.id))

    filas = "".join(
        f"<tr><td>{h.numero}</td><td>{h.descripcion_a}</td>"
        f"<td><input name='h{h.numero}'></td></tr>" for h in e.hilos)
    cuerpo = f"""
    <p><b>{e.anillo} | {e.nombre}</b> · {e.capacidad} hilos</p>
    <form method="post">
      <label>Origen B</label><input name="origen_b" value="{e.origen_b}" required>
      <table><tr><th>Fibra</th><th>Extremo A (referencia)</th><th>Descripción Extremo B</th></tr>{filas}</table>
      <button class="btn" style="background:#059669">Guardar y Validar</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM, titulo="Completar Extremo B", cuerpo=cuerpo)


@app.route("/enlace/<id>/editar", methods=["GET", "POST"])
def editar(id):
    e = Enlace.query.get_or_404(id)
    if request.method == "POST":
        e.nombre = request.form.get("nombre", e.nombre)
        e.origen_a = request.form.get("origen_a", e.origen_a)
        e.origen_b = request.form.get("origen_b", e.origen_b)
        for h in e.hilos:
            h.descripcion_a = request.form.get(f"a{h.numero}", h.descripcion_a)
            h.descripcion_b = request.form.get(f"b{h.numero}", h.descripcion_b)
        db.session.commit()
        return redirect(url_for("dashboard"))

    filas = "".join(
        f"<tr><td>{h.numero}</td>"
        f"<td><input name='a{h.numero}' value=\"{h.descripcion_a}\"></td>"
        f"<td><input name='b{h.numero}' value=\"{h.descripcion_b}\"></td></tr>"
        for h in e.hilos)
    cuerpo = f"""
    <form method="post">
      <label>Nombre</label><input name="nombre" value="{e.nombre}">
      <label>Origen A</label><input name="origen_a" value="{e.origen_a}">
      <label>Origen B</label><input name="origen_b" value="{e.origen_b}">
      <table><tr><th>Fibra</th><th>Extremo A</th><th>Extremo B</th></tr>{filas}</table>
      <button class="btn">Guardar cambios</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM, titulo="Editar Enlace", cuerpo=cuerpo)


@app.route("/enlace/<id>/ver")
def ver(id):
    e = Enlace.query.get_or_404(id)
    difs = sum(1 for h in e.hilos if not h.coincide)
    filas = "".join(
        f"<tr class='{'rojo' if not h.coincide else ''}'>"
        f"<td>{h.numero}</td><td>{h.descripcion_a}</td>"
        f"<td>{h.descripcion_b}</td><td>{'Sí' if h.coincide else 'No'}</td></tr>"
        for h in e.hilos)
    aviso = (f"<p style='color:#b45309'>⚠ {difs} hilo(s) con diferencias.</p>"
             if difs else "<p style='color:#059669'>✔ Todos los hilos coinciden.</p>")
    cuerpo = f"""
    <p><b>{e.anillo} | {e.nombre}</b> · {e.capacidad} hilos ·
       {e.origen_a} ⇄ {e.origen_b} · Estado: {e.estado}</p>
    {aviso}
    <table><tr><th>Fibra</th><th>{e.origen_a or 'A'}</th>
           <th>{e.origen_b or 'B'}</th><th>Coincide</th></tr>{filas}</table>
    <a class="btn btn-borde-verde" style="background:#059669"
       href="{url_for('exportar', id=e.id)}">Exportar Excel</a>
    <a class="btn" style="background:#6b7280" href="/">Volver</a>"""
    return render_template_string(TPL_FORM, titulo="Reporte Comparativo", cuerpo=cuerpo)


@app.route("/enlace/<id>/export.xlsx")
def exportar(id):
    e = Enlace.query.get_or_404(id)
    wb = Workbook(); ws = wb.active; ws.title = "REPORTE"

    azul = PatternFill("solid", start_color="1F4E79")
    rojo = PatternFill("solid", start_color="F8CBAD")
    thin = Side(style="thin", color="BFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{e.anillo} | {e.nombre}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = (f"Capacidad: {e.capacidad} · {e.origen_a} ⇄ {e.origen_b} · "
                f"Estado: {e.estado} · Creado: {e.creado:%d/%m/%Y %H:%M}")
    ws["A2"].font = Font(size=9, color="595959")

    ws.append([]); ws.append(["FIBRA", f"DESCRIPCIÓN {e.origen_a or 'A'}",
                              f"DESCRIPCIÓN {e.origen_b or 'B'}", "COINCIDE"])
    hdr = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(hdr, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde

    for h in e.hilos:
        ws.append([h.numero, h.descripcion_a, h.descripcion_b,
                   "Sí" if h.coincide else "No"])
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(r, c).border = borde
        if not h.coincide:
            ws.cell(r, 4).fill = rojo

    for col, w in zip("ABCD", [8, 50, 50, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hdr+1}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{e.nombre.replace(' ', '_')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True)
