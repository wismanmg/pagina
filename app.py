# -*- coding: utf-8 -*-
"""
VALIDACIÓN DE ENLACES — Reporte de Hilos (v2: Extremo A y Extremo B)
=====================================================================
Aplicación completa en UN SOLO archivo Python (Flask + SQLite).

FLUJO v2:
  1. "+ Nuevo Enlace"  → se crea solo la FICHA (anillo, nombre, capacidad).
  2. "Completar A"     → se registra Origen A + descripción de cada hilo (lado A).
  3. "Completar B"     → se registra Origen B + descripción de cada hilo (lado B).
  4. Cuando A y B están completos → el enlace pasa a VALIDADO automáticamente
     y aparece el reporte comparativo hilo a hilo (Coincide Sí/No).

CÓMO EJECUTAR:
    pip install flask flask-sqlalchemy openpyxl
    python reporte_hilos.py
    → abrir http://127.0.0.1:5000
"""

import io
import uuid
from datetime import datetime

from flask import (Flask, request, redirect, url_for, send_file,
                   render_template_string)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///enlaces.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

# ============================================================
# 1) MODELOS
# ============================================================
class Enlace(db.Model):
    id = db.Column(db.String(36), primary_key=True,
                   default=lambda: str(uuid.uuid4()))
    anillo = db.Column(db.String(30), nullable=False)
    nombre = db.Column(db.String(120), nullable=False)
    capacidad = db.Column(db.Integer, nullable=False, default=24)
    tipo_cable = db.Column(db.String(60), default="CABLE DE ENLACE SM")
    longitud = db.Column(db.String(40), default="")
    origen_a = db.Column(db.String(80), default="")
    sala_a = db.Column(db.String(80), default="")
    rack_a = db.Column(db.String(40), default="")
    posicion_a = db.Column(db.String(40), default="")
    origen_b = db.Column(db.String(80), default="")
    sala_b = db.Column(db.String(80), default="")
    rack_b = db.Column(db.String(40), default="")
    posicion_b = db.Column(db.String(40), default="")
    a_completo = db.Column(db.Boolean, default=False)   # ← lado A registrado
    b_completo = db.Column(db.Boolean, default=False)   # ← lado B registrado
    creado = db.Column(db.DateTime, default=datetime.utcnow)
    hilos = db.relationship("Hilo", backref="enlace",
                            cascade="all, delete-orphan", order_by="Hilo.numero")

    @property
    def estado(self):
        """VALIDADO solo cuando AMBOS extremos están completos."""
        return "VALIDADO" if (self.a_completo and self.b_completo) else "INCOMPLETO"

    @property
    def pendiente_texto(self):
        if not self.a_completo and not self.b_completo:
            return "Falta Extremo A y Extremo B"
        if not self.a_completo:
            return "Falta Extremo A"
        if not self.b_completo:
            return "Falta Extremo B"
        return ""


class Hilo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    enlace_id = db.Column(db.String(36), db.ForeignKey("enlace.id"))
    numero = db.Column(db.Integer)
    descripcion_a = db.Column(db.Text, default="")
    descripcion_b = db.Column(db.Text, default="")

    @property
    def coincide(self):
        na = " ".join((self.descripcion_a or "").upper().split())
        nb = " ".join((self.descripcion_b or "").upper().split())
        return na == nb


# ============================================================
# 2) DATOS INICIALES (tu dashboard: 27 pendientes con A hecho + 3 validados)
#    (anillo, nombre, origen_a, origen_b, capacidad, fecha, a_ok, b_ok)
# ============================================================
SEED = [
    ("ANILLO ESTE",  "SANTA ANITA - LA MOLINA",   "SANTA ANITA", "LA MOLINA",   48, "2026-04-21 17:18:46", True, False),
    ("ANILLO SUR",   "LURIN - TELEPUERTO - 96SM", "LURIN",       "96SM",        96, "2026-04-18 20:10:17", True, False),
    ("ANILLO ESTE",  "AVIACION - HIGUERETA",      "AVIACION",    "HIGUERETA",   96, "2026-03-12 02:40:42", True, False),
    ("ANILLO ESTE",  "CAMACHO - MONTERICO",       "MONTERICO",   "CAMACHO",     48, "2026-03-08 00:52:18", True, False),
    ("ANILLO ESTE",  "SAN LUIS - MOTERRICO",      "MONTERICO",   "SAN LUIS",    48, "2026-03-07 02:30:07", True, False),
    ("ANILLO ESTE",  "HIGUERETA - SAN LUIS",      "HIGUERETA",   "SAN LUIS",    48, "2026-02-05 19:00:47", True, False),
    ("ANILLO ESTE",  "HIGUERETA - TELEPUERTO",    "HIGUERETA",   "TELEPUERTO",  48, "2026-02-05 14:13:29", True, False),
    ("ANILLO SUR",   "OCHARAN - HIGUERETA",       "OCHARAN",     "HIGUERETA",   96, "2026-02-04 21:05:33", True, False),
    ("ANILLO OESTE", "SAN MIGUEL - SAN FELIPE 1", "POP SAN MIGUEL", "SAN FELIPE 1", 48, "2026-01-21 20:40:26", True, False),
    ("ANILLO NORTE", "COTABAMBAS - ZARATE",       "COTABAMBAS",  "ZARATE",      96, "2026-01-20 22:26:54", True, False),
    ("ANILLO NORTE", "COLONIAL - COTABAMBAS",     "COLONIAL",    "COTABAMBAS",  48, "2026-01-20 17:16:31", True, False),
    ("ANILLO OESTE", "AEROPUERTO - CALLAO",       "AEROPUERTO",  "CALLAO",      96, "2026-01-19 20:28:32", True, False),
    ("ANILLO SUR",   "AERIPUETO - LA MILLA",      "AERIPUETO",   "LA MILLA",    24, "2026-01-19 17:26:43", True, False),
    ("ANILLO SUR",   "AEROPUERTO - PAMPA LIBRE",  "AEROPUERTO",  "PAMPA LIBRE", 24, "2026-01-19 14:28:30", True, False),
    ("ANILLO NORTE", "AEROPUERTO - COLONIAL",     "AEROPUERTO",  "COLONIAL",    48, "2026-01-15 16:10:38", True, False),
    ("ANILLO SUR",   "AEROPUERTO - POLO 1",       "AEROPUERTO",  "POLO 1",      24, "2026-01-15 02:07:20", True, False),
    ("ANILLO OESTE", "LA PUNTA - SAN MIGUEL",     "LA PUNTA",    "SAN MIGUEL",  48, "2026-01-14 23:20:24", True, False),
    ("ANILLO OESTE", "CALLO - LA PUNTA",          "CALLAO",      "LA PUNTA",    48, "2026-01-14 22:34:45", True, False),
    ("ANILLO SUR",   "TELEPUERTO - ASIA",         "TELEPUERTO",  "ASIA",        24, "2026-01-14 02:16:00", True, False),
    ("ANILLO NORTE", "SANTA LUZMILA - LOS OLIVOS","SANTA LUZMILA","LOS OLIVOS", 48, "2026-01-12 18:38:00", True, False),
    ("ANILLO NORTE", "AEROPUERTO - LOS OLIVOS",   "AEROPUERTO",  "LOS OLIVOS",  48, "2026-01-12 16:40:00", True, False),
    ("ANILLO ESTE",  "ZARATE - SANTA ANITA",      "ZARATE",      "SANTA ANITA", 48, "2026-01-10 19:30:00", True, False),
    ("ANILLO NORTE", "ZARATE - INGENIERIA",       "ZARATE",      "INGENIERIA",  48, "2026-01-10 17:58:00", True, False),
    ("ANILLO SUR",   "TELPUERTO -POLO 2",         "POLO 2",      "VILLA EL SAVADOR", 24, "2026-01-09 23:12:00", True, False),
    ("ANILLO SUR",   "POLO 2 - POLO 1",           "P0L0 1",      "POLO 2",      24, "2026-01-09 18:01:00", True, False),
    ("ANILLO SUR",   "CHORRILLOS - OCHARAN",      "CHORRILLOS",  "OCHARAN",     96, "2026-01-08 18:39:00", True, False),
    ("ANILLO SUR",   "SAN JUAN - TELPUERTO",      "SAN JUAN",    "TELPUERTO",   96, "2026-01-07 22:31:00", True, False),
    # ---- Validados (A y B completos) ----
    ("ANILLO SUR",   "LURIN - TELEPUERTO - 48SM", "TELEPUERTO",  "LURIN",       48, "2026-01-14 03:15:26", True, True),
    ("ANILLO SUR",   "SAN JUAN - POLO 2",         "SAN JUAN",    "POLO 2",      96, "2026-01-08 00:14:39", True, True),
    ("ANILLO SUR",   "SAN JUAN - CHORRILLOS",     "SAN JUAN",    "CHORRILLOS",  48, "2026-01-07 20:04:48", True, True),
]


CATALOGO_ENLACES = sorted({s[1] for s in SEED})
TIPOS_CABLE = ["CABLE DE ENLACE SM", "ADSS", "OPGW", "FIGURA 8", "ARMADO", "DUCTO", "OTRO"]


def seed_inicial():
    if Enlace.query.count() > 0:
        return
    for anillo, nombre, a, b, cap, fecha, a_ok, b_ok in SEED:
        e = Enlace(anillo=anillo, nombre=nombre, capacidad=cap,
                   origen_a=a, origen_b=b, a_completo=a_ok, b_completo=b_ok,
                   creado=datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S"))
        for n in range(1, cap + 1):
            e.hilos.append(Hilo(numero=n,
                                descripcion_a="LIBRE" if a_ok else "",
                                descripcion_b="LIBRE" if b_ok else ""))
        db.session.add(e)
    db.session.commit()


with app.app_context():
    db.create_all()
    seed_inicial()

# ============================================================
# 3) PLANTILLAS
# ============================================================
TPL_DASH = """
<!doctype html><html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard - Validación</title>
<style>
 body{margin:0;font-family:'Segoe UI',system-ui,sans-serif;background:#fff;color:#1f2937}
 header{display:flex;align-items:center;justify-content:space-between;
        padding:14px 24px;border-bottom:1px solid #e5e7eb}
 header h1{font-size:1.15rem;margin:0}
 .btn{display:inline-block;padding:6px 12px;border-radius:6px;border:1px solid #d1d5db;
      background:#fff;color:#374151;text-decoration:none;font-size:.78rem;cursor:pointer}
 .btn-azul{background:#2563eb;border-color:#2563eb;color:#fff}
 .btn-borde-azul{border-color:#2563eb;color:#2563eb}
 .btn-borde-verde{border-color:#059669;color:#059669}
 .btn-borde-naranja{border-color:#d97706;color:#d97706}
 .grid{display:grid;grid-template-columns:1fr 1fr;gap:24px;padding:24px;
       max-width:1550px;margin:0 auto;align-items:start}
 .card{background:#fff;border:1px solid #e5e7eb;border-radius:12px;
       box-shadow:0 1px 4px rgba(0,0,0,.06);padding:18px}
 .card h2{font-size:1rem;margin:0 0 2px;display:flex;justify-content:space-between}
 .badge{background:#dbeafe;color:#1d4ed8;border-radius:10px;padding:1px 9px;font-size:.75rem}
 .sub{color:#6b7280;font-size:.75rem;margin:0 0 12px}
 table{width:100%;border-collapse:collapse;font-size:.8rem}
 th{text-align:left;padding:6px 4px;border-bottom:2px solid #e5e7eb}
 td{padding:9px 4px;border-bottom:1px solid #f3f4f6;vertical-align:middle}
 .tramo{color:#9ca3af;font-size:.72rem}
 .falta{color:#d97706;font-size:.7rem;font-weight:600}
 .acc{display:flex;gap:5px;flex-wrap:wrap}
</style></head><body>
<header>
  <h1>🧵 Validación de Enlaces</h1>
  <div>
    <a class="btn btn-borde-verde" style="margin-right:6px"
       href="{{ url_for('reporte_avance') }}">📊 Reporte de Avance</a>
    <a class="btn btn-azul" href="{{ url_for('nuevo') }}">+ Nuevo Enlace</a>
  </div>
</header>
<div class="grid">

  <div class="card">
    <h2>Pendientes (Incompletos) <span class="badge">{{ pend|length }}</span></h2>
    <p class="sub">Enlaces con Extremo A y/o Extremo B pendientes de completar.</p>
    <table>
      <tr><th>Enlace</th><th>Cap.</th><th>Creado</th><th>Acciones</th></tr>
      {% for e in pend %}
      <tr>
        <td><strong>{{ e.anillo }} | {{ e.nombre }}</strong><br>
            <span class="tramo">{{ e.origen_a or '?' }} → {{ e.origen_b or '?' }}</span><br>
            <span class="falta">{{ e.pendiente_texto }}</span></td>
        <td>{{ e.capacidad }}</td>
        <td>{{ e.creado.strftime('%Y-%m-%d %H:%M:%S') }}</td>
        <td class="acc">
          {% if not e.a_completo %}
            <a class="btn btn-borde-naranja" href="{{ url_for('completar', id=e.id, lado='a') }}">Completar A</a>
          {% endif %}
          {% if not e.b_completo %}
            <a class="btn btn-borde-azul" href="{{ url_for('completar', id=e.id, lado='b') }}">Completar B</a>
          {% endif %}
          <a class="btn" href="{{ url_for('editar', id=e.id) }}">Editar</a>
          <a class="btn btn-borde-verde" href="{{ url_for('exportar', id=e.id) }}">Exportar</a>
        </td>
      </tr>
      {% endfor %}
    </table>
  </div>

  <div class="card">
    <h2>Validados <span class="badge">{{ valid|length }}</span></h2>
    <p class="sub">Enlaces con Extremo A y B completados (reporte comparativo).</p>
    <table>
      <tr><th>Enlace</th><th>Cap.</th><th>Creado</th><th>Acciones</th></tr>
      {% for e in valid %}
      <tr>
        <td><strong>{{ e.anillo }} | {{ e.nombre }}</strong><br>
            <span class="tramo">{{ e.origen_a }} ⇄ {{ e.origen_b }}</span></td>
        <td>{{ e.capacidad }}</td>
        <td>{{ e.creado.strftime('%Y-%m-%d %H:%M:%S') }}</td>
        <td class="acc">
          <a class="btn" href="{{ url_for('ver', id=e.id) }}">Ver</a>
          <a class="btn" href="{{ url_for('editar', id=e.id) }}">Editar</a>
          <a class="btn btn-borde-verde" href="{{ url_for('exportar', id=e.id) }}">Exportar</a>
        </td>
      </tr>
      {% endfor %}
    </table>
  </div>

</div></body></html>
"""

TPL_FORM = """
<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ titulo }}</title>
<style>
 body{font-family:'Segoe UI',sans-serif;max-width:940px;margin:24px auto;color:#1f2937;padding:0 12px}
 label{display:block;font-size:.8rem;color:#6b7280;margin-top:10px}
 input,select,textarea{width:100%;padding:6px;border:1px solid #d1d5db;border-radius:6px;box-sizing:border-box}
 table{width:100%;border-collapse:collapse;font-size:.8rem;margin-top:14px}
 th,td{border:1px solid #e5e7eb;padding:5px}
 th{background:#f3f4f6}
 .btn{display:inline-block;margin-top:16px;padding:8px 16px;border-radius:6px;
      background:#2563eb;color:#fff;border:0;cursor:pointer;text-decoration:none}
 .rojo{background:#fee2e2}
 .ref{color:#6b7280}
</style></head><body>
<h2>{{ titulo }}</h2>
{{ cuerpo|safe }}
</body></html>
"""

# ============================================================
# 4) RUTAS
# ============================================================
@app.route("/")
def dashboard():
    todos = Enlace.query.order_by(Enlace.creado.desc()).all()
    pend = [e for e in todos if e.estado == "INCOMPLETO"]
    valid = [e for e in todos if e.estado == "VALIDADO"]
    return render_template_string(TPL_DASH, pend=pend, valid=valid)


@app.route("/nuevo", methods=["GET", "POST"])
def nuevo():
    """Crea solo la FICHA del enlace. A y B se completan después."""
    if request.method == "POST":
        cap = int(request.form.get("capacidad") or 24)
        e = Enlace(anillo=request.form.get("anillo", ""),
                   nombre=request.form.get("nombre", "").strip(),
                   capacidad=cap,
                   origen_a=request.form.get("origen_a", "").strip(),
                   origen_b=request.form.get("origen_b", "").strip())
        for n in range(1, cap + 1):
            e.hilos.append(Hilo(numero=n))
        db.session.add(e)
        db.session.commit()
        return redirect(url_for("dashboard"))

    cuerpo = """
    <p class="ref">Se crea la ficha del enlace. Luego usa <b>Completar A</b> y
    <b>Completar B</b> desde el dashboard para registrar cada extremo.</p>
    <form method="post">
      <label>Anillo</label>
      <select name="anillo">
        <option>ANILLO NORTE</option><option>ANILLO SUR</option>
        <option>ANILLO ESTE</option><option>ANILLO OESTE</option>
      </select>
      <label>Nombre del enlace (SITIO A - SITIO B)</label><input name="nombre" required>
      <label>Capacidad (hilos)</label>
      <select name="capacidad"><option>24</option><option selected>48</option><option>96</option></select>
      <label>Origen A (sugerido)</label><input name="origen_a">
      <label>Origen B (sugerido)</label><input name="origen_b">
      <button class="btn">Crear enlace</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM, titulo="Nuevo Enlace", cuerpo=cuerpo)


@app.route("/enlace/<id>/completar/<lado>", methods=["GET", "POST"])
def completar(id, lado):
    """Formulario para completar el Extremo A o el Extremo B."""
    if lado not in ("a", "b"):
        return redirect(url_for("dashboard"))
    e = Enlace.query.get_or_404(id)
    es_a = (lado == "a")

    if request.method == "POST":
        origen = request.form.get("origen", "").strip()
        sala = request.form.get("sala", "").strip()
        rack = request.form.get("rack", "").strip()
        pos = request.form.get("posicion", "").strip()
        if es_a:
            e.origen_a = origen or e.origen_a
            e.sala_a, e.rack_a, e.posicion_a = sala, rack, pos
            for h in e.hilos:
                h.descripcion_a = request.form.get(f"h{h.numero}", "").strip()
            e.a_completo = True
        else:
            e.origen_b = origen or e.origen_b
            e.sala_b, e.rack_b, e.posicion_b = sala, rack, pos
            for h in e.hilos:
                h.descripcion_b = request.form.get(f"h{h.numero}", "").strip()
            e.b_completo = True
        db.session.commit()
        # Si con esto quedó VALIDADO, ir directo al comparativo
        if e.estado == "VALIDADO":
            return redirect(url_for("ver", id=e.id))
        return redirect(url_for("dashboard"))

    # --- GET: armar formulario ---
    letra = "A" if es_a else "B"
    origen_actual = e.origen_a if es_a else e.origen_b
    # el otro lado se muestra como referencia si ya existe
    otro_ok = e.b_completo if es_a else e.a_completo
    otro_nom = e.origen_b if es_a else e.origen_a

    filas = []
    for h in e.hilos:
        ref = h.descripcion_b if es_a else h.descripcion_a
        val = h.descripcion_a if es_a else h.descripcion_b
        celda_ref = (f"<td class='ref'>{ref}</td>" if otro_ok else "")
        filas.append(f"<tr><td>{h.numero}</td>{celda_ref}"
                     f"<td><input name='h{h.numero}' value=\"{val}\" "
                     f"placeholder='LIBRE / ODF / SIN PIGTAIL...'></td></tr>")
    th_ref = (f"<th>Extremo {'B' if es_a else 'A'} — {otro_nom} (referencia)</th>"
              if otro_ok else "")

    cuerpo = f"""
    <p><b>{e.anillo} | {e.nombre}</b> · {e.capacidad} hilos ·
       Estado actual: {e.estado} ({e.pendiente_texto or 'completo'})</p>
    <form method="post">
      <label>Origen (Sitio {letra})</label>
      <input name="origen" value="{origen_actual}" required>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px">
        <div><label>Sala {letra}</label>
          <input name="sala" value="{(e.sala_a if es_a else e.sala_b) or ''}"></div>
        <div><label>Rack {letra}</label>
          <input name="rack" value="{(e.rack_a if es_a else e.rack_b) or ''}"></div>
        <div><label>Posición {letra}</label>
          <input name="posicion" value="{(e.posicion_a if es_a else e.posicion_b) or ''}"></div>
      </div>
      <table>
        <tr><th>Fibra</th>{th_ref}<th>Descripción Extremo {letra}</th></tr>
        {''.join(filas)}
      </table>
      <button class="btn" style="background:{'#d97706' if es_a else '#2563eb'}">
        Guardar Extremo {letra}</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM,
                                  titulo=f"Completar Extremo {letra}", cuerpo=cuerpo)


@app.route("/enlace/<id>/editar", methods=["GET", "POST"])
def editar(id):
    e = Enlace.query.get_or_404(id)
    if request.method == "POST":
        # --- Nombre: select "mantener" o manual ---
        sel = request.form.get("nombre_sel", "").strip()
        manual = request.form.get("nombre_manual", "").strip()
        if manual:
            e.nombre = manual
        elif sel and sel != "__actual__":
            e.nombre = sel

        e.tipo_cable = request.form.get("tipo_cable", e.tipo_cable).strip()
        e.longitud = request.form.get("longitud", e.longitud).strip()

        # --- Capacidad: si reduce, se eliminan hilos mayores a N; si aumenta, se crean ---
        try:
            nueva_cap = int(request.form.get("capacidad") or e.capacidad)
        except ValueError:
            nueva_cap = e.capacidad
        if nueva_cap != e.capacidad and nueva_cap > 0:
            if nueva_cap < e.capacidad:
                for h in list(e.hilos):
                    if h.numero > nueva_cap:
                        db.session.delete(h)
            else:
                for n in range(e.capacidad + 1, nueva_cap + 1):
                    e.hilos.append(Hilo(numero=n))
            e.capacidad = nueva_cap

        # --- Extremo A ---
        e.origen_a = request.form.get("origen_a", e.origen_a).strip()
        e.sala_a = request.form.get("sala_a", e.sala_a).strip()
        e.rack_a = request.form.get("rack_a", e.rack_a).strip()
        e.posicion_a = request.form.get("posicion_a", e.posicion_a).strip()

        # --- Extremo B (nota: completarlo aqui no marca b_completo;
        #     eso se hace desde "Completar B") ---
        e.origen_b = request.form.get("origen_b", e.origen_b).strip()
        e.sala_b = request.form.get("sala_b", e.sala_b).strip()
        e.rack_b = request.form.get("rack_b", e.rack_b).strip()
        e.posicion_b = request.form.get("posicion_b", e.posicion_b).strip()

        # --- Descripciones de hilos (si vienen en el form) ---
        for h in e.hilos:
            if f"a{h.numero}" in request.form:
                h.descripcion_a = request.form.get(f"a{h.numero}", h.descripcion_a)
            if f"b{h.numero}" in request.form:
                h.descripcion_b = request.form.get(f"b{h.numero}", h.descripcion_b)

        db.session.commit()
        return redirect(url_for("dashboard"))

    # ---------- GET: formulario con el diseno de la app original ----------
    opciones = "".join(
        f"<option value=\"{n}\" {'selected' if n == e.nombre else ''}>{n}</option>"
        for n in CATALOGO_ENLACES)
    tipos = "".join(
        f"<option {'selected' if t == (e.tipo_cable or '') else ''}>{t}</option>"
        for t in TIPOS_CABLE)
    filas = "".join(
        f"<tr><td>{h.numero}</td>"
        f"<td><input name='a{h.numero}' value=\"{h.descripcion_a}\"></td>"
        f"<td><input name='b{h.numero}' value=\"{h.descripcion_b}\"></td></tr>"
        for h in e.hilos)

    cuerpo = f"""
    <style>
      .top{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px}}
      .top .acciones a{{margin-left:6px}}
      .fila3{{display:grid;grid-template-columns:2fr 1.2fr 1fr;gap:18px}}
      .dosc{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:18px}}
      .cardx{{background:#f9fafb;border:1px solid #e5e7eb;border-radius:10px;padding:14px}}
      .cardx h4{{margin:0 0 8px;font-size:.9rem}}
      .g3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}}
      .nota{{color:#6b7280;font-size:.72rem;margin-top:6px}}
      .btn-borde{{background:#fff;color:#374151;border:1px solid #d1d5db}}
      .btn-verde{{background:#fff;color:#059669;border:1px solid #059669}}
    </style>
    <div class="top">
      <div>
        <p class="ref" style="margin:0">Puedes corregir cabecera, Extremos A/B y descripciones.</p>
      </div>
      <div class="acciones">
        <a class="btn btn-borde" href="/">&larr; Volver</a>
        <a class="btn btn-borde" href="{url_for('ver', id=e.id)}">Ver reporte</a>
        <a class="btn btn-verde" href="{url_for('exportar', id=e.id)}">Exportar</a>
      </div>
    </div>
    <form method="post">
      <div class="fila3">
        <div>
          <label>Nombre del Enlace</label>
          <select name="nombre_sel">
            <option value="__actual__" selected>(Mantener actual)</option>
            {opciones}
          </select>
          <input name="nombre_manual" placeholder="(Opcional) Nombre manual del enlace"
                 style="margin-top:6px">
          <div class="nota">Actual: <b>{e.anillo} | {e.nombre}</b></div>
        </div>
        <div>
          <label>Tipo de Cable</label>
          <select name="tipo_cable">{tipos}</select>
        </div>
        <div>
          <label>Capacidad (N hilos)</label>
          <input name="capacidad" type="number" min="1" value="{e.capacidad}">
          <div class="nota">Si reduces capacidad, se eliminan hilos mayores a N.</div>
        </div>
      </div>

      <label style="margin-top:14px">Longitud Tramo Total</label>
      <input name="longitud" value="{e.longitud or ''}" style="max-width:320px">

      <div class="dosc">
        <div class="cardx">
          <h4>Extremo A</h4>
          <div class="g3">
            <div><label>Origen A</label><input name="origen_a" value="{e.origen_a}"></div>
            <div><label>Sala A</label><input name="sala_a" value="{e.sala_a or ''}"></div>
            <div><label>Rack A</label><input name="rack_a" value="{e.rack_a or ''}"></div>
          </div>
          <div style="max-width:32%"><label>Posici&oacute;n A</label>
            <input name="posicion_a" value="{e.posicion_a or ''}"></div>
        </div>
        <div class="cardx">
          <h4>Extremo B</h4>
          <div class="g3">
            <div><label>Origen B</label><input name="origen_b" value="{e.origen_b}"></div>
            <div><label>Sala B</label><input name="sala_b" value="{e.sala_b or ''}"></div>
            <div><label>Rack B</label><input name="rack_b" value="{e.rack_b or ''}"></div>
          </div>
          <div style="max-width:32%"><label>Posici&oacute;n B</label>
            <input name="posicion_b" value="{e.posicion_b or ''}"></div>
          <div class="nota">Nota: si completas B por aqu&iacute;, igual puedes ir a
            "Completar B" si est&aacute; pendiente.</div>
        </div>
      </div>

      <h4 style="margin-top:20px">Descripciones de hilos</h4>
      <table><tr><th>Fibra</th><th>Extremo A</th><th>Extremo B</th></tr>{filas}</table>
      <button class="btn">Guardar cambios</button>
      <a class="btn" style="background:#6b7280" href="/">Cancelar</a>
    </form>"""
    return render_template_string(TPL_FORM, titulo="Editar Enlace", cuerpo=cuerpo)


@app.route("/enlace/<id>/ver")
def ver(id):
    e = Enlace.query.get_or_404(id)
    difs = sum(1 for h in e.hilos if not h.coincide)
    filas = "".join(
        f"<tr class='{'rojo' if not h.coincide else ''}'>"
        f"<td>{h.numero}</td><td>{h.descripcion_a}</td>"
        f"<td>{h.descripcion_b}</td><td>{'Sí' if h.coincide else 'No'}</td></tr>"
        for h in e.hilos)
    aviso = (f"<p style='color:#b45309'>⚠ {difs} hilo(s) con diferencias entre A y B.</p>"
             if difs else "<p style='color:#059669'>✔ Todos los hilos coinciden.</p>")
    cuerpo = f"""
    <p><b>{e.anillo} | {e.nombre}</b> · {e.capacidad} hilos ·
       {e.origen_a} ⇄ {e.origen_b} · Estado: {e.estado}</p>
    {aviso}
    <table><tr><th>Fibra</th><th>{e.origen_a or 'A'}</th>
           <th>{e.origen_b or 'B'}</th><th>Coincide</th></tr>{filas}</table>
    <a class="btn" style="background:#059669"
       href="{url_for('exportar', id=e.id)}">Exportar Excel</a>
    <a class="btn" style="background:#6b7280" href="/">Volver</a>"""
    return render_template_string(TPL_FORM, titulo="Reporte Comparativo", cuerpo=cuerpo)


@app.route("/enlace/<id>/export.xlsx")
def exportar(id):
    e = Enlace.query.get_or_404(id)
    wb = Workbook(); ws = wb.active; ws.title = "REPORTE"

    azul = PatternFill("solid", start_color="1F4E79")
    rojo = PatternFill("solid", start_color="F8CBAD")
    thin = Side(style="thin", color="BFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{e.anillo} | {e.nombre}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = (f"Capacidad: {e.capacidad} · Tipo: {e.tipo_cable or '-'} · "
                f"Longitud: {e.longitud or '-'} · Estado: {e.estado} · "
                f"Creado: {e.creado:%d/%m/%Y %H:%M}")
    ws["A3"] = (f"EXTREMO A: {e.origen_a or '?'} | Sala: {e.sala_a or '-'} | "
                f"Rack: {e.rack_a or '-'} | Posición: {e.posicion_a or '-'}")
    ws["A4"] = (f"EXTREMO B: {e.origen_b or '?'} | Sala: {e.sala_b or '-'} | "
                f"Rack: {e.rack_b or '-'} | Posición: {e.posicion_b or '-'}")
    ws["A3"].font = Font(size=9, color="595959")
    ws["A4"].font = Font(size=9, color="595959")
    ws["A2"].font = Font(size=9, color="595959")

    ws.append([]); ws.append(["FIBRA", f"DESCRIPCIÓN {e.origen_a or 'A'}",
                              f"DESCRIPCIÓN {e.origen_b or 'B'}", "COINCIDE"])
    hdr = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(hdr, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde

    for h in e.hilos:
        ws.append([h.numero, h.descripcion_a, h.descripcion_b,
                   "Sí" if h.coincide else "No"])
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(r, c).border = borde
        if not h.coincide:
            ws.cell(r, 4).fill = rojo

    for col, w in zip("ABCD", [8, 50, 50, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hdr+1}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{e.nombre.replace(' ', '_')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")



# ============================================================
# 5) REPORTE DE AVANCE (Excel consolidado generado en vivo)
# ============================================================
@app.route("/reporte-avance.xlsx")
def reporte_avance():
    """Excel con RESUMEN (KPIs, por anillo, por capacidad) + detalle
    de pendientes y validados, calculado con los datos actuales."""
    todos = Enlace.query.order_by(Enlace.creado.asc()).all()
    pend = [e for e in todos if e.estado == "INCOMPLETO"]
    valid = [e for e in todos if e.estado == "VALIDADO"]
    ahora = datetime.now()

    azul = PatternFill("solid", start_color="1F4E79")
    azul2 = PatternFill("solid", start_color="2E75B6")
    gris = PatternFill("solid", start_color="F2F2F2")
    verde = PatternFill("solid", start_color="70AD47")
    thin = Side(style="thin", color="BFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_row(ws, fila, textos, fill=azul):
        for i, t in enumerate(textos, 1):
            c = ws.cell(fila, i, t)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde

    wb = Workbook()

    # ---------- HOJA RESUMEN ----------
    rs = wb.active
    rs.title = "RESUMEN"
    rs["A1"] = "REPORTE DE AVANCE — VALIDACIÓN DE ENLACES"
    rs["A1"].font = Font(bold=True, size=14, color="1F4E79")
    rs["A2"] = f"Generado: {ahora:%d/%m/%Y %H:%M}"
    rs["A2"].font = Font(size=9, color="595959")

    total = len(todos)
    hilos_pend = sum(e.capacidad for e in pend)
    hilos_ok = sum(e.capacidad for e in valid)
    dias = lambda e: (ahora - e.creado).days
    kpis = [
        ("Total de enlaces", total),
        ("Enlaces validados (A y B)", len(valid)),
        ("Enlaces pendientes", len(pend)),
        ("% de avance de validación",
         f"{(len(valid)/total*100):.1f}%" if total else "0%"),
        ("Hilos validados", hilos_ok),
        ("Hilos pendientes", hilos_pend),
        ("Pendiente más antiguo (días)", max((dias(e) for e in pend), default=0)),
        ("Antigüedad promedio pendientes (días)",
         round(sum(dias(e) for e in pend)/len(pend)) if pend else 0),
        ("Pendientes solo por Extremo B",
         sum(1 for e in pend if e.a_completo and not e.b_completo)),
        ("Pendientes solo por Extremo A",
         sum(1 for e in pend if e.b_completo and not e.a_completo)),
        ("Pendientes por ambos extremos",
         sum(1 for e in pend if not e.a_completo and not e.b_completo)),
    ]
    rs["A4"] = "INDICADORES GENERALES"
    rs["A4"].font = Font(bold=True, size=11, color="1F4E79")
    hdr_row(rs, 5, ["INDICADOR", "VALOR"], azul2)
    for i, (k, v) in enumerate(kpis):
        r = 6 + i
        rs.cell(r, 1, k).border = borde
        c = rs.cell(r, 2, v)
        c.border = borde
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Avance por anillo
    anillos = ["ANILLO NORTE", "ANILLO SUR", "ANILLO ESTE", "ANILLO OESTE"]
    rs["D4"] = "AVANCE POR ANILLO"
    rs["D4"].font = Font(bold=True, size=11, color="1F4E79")
    hdr_row_cells = ["ANILLO", "PENDIENTES", "VALIDADOS", "% AVANCE"]
    for i, t in enumerate(hdr_row_cells):
        c = rs.cell(5, 4 + i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = azul2
        c.alignment = Alignment(horizontal="center")
        c.border = borde
    for i, an in enumerate(anillos):
        r = 6 + i
        np_ = sum(1 for e in pend if e.anillo == an)
        nv = sum(1 for e in valid if e.anillo == an)
        pct = f"{(nv/(np_+nv)*100):.0f}%" if (np_ + nv) else "-"
        for j, v in enumerate([an, np_, nv, pct]):
            c = rs.cell(r, 4 + j, v)
            c.border = borde
            if j:
                c.alignment = Alignment(horizontal="center")

    # Por capacidad
    rs["D12"] = "PENDIENTES POR CAPACIDAD"
    rs["D12"].font = Font(bold=True, size=11, color="1F4E79")
    for i, t in enumerate(["CAPACIDAD", "ENLACES", "HILOS"]):
        c = rs.cell(13, 4 + i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = azul2
        c.alignment = Alignment(horizontal="center")
        c.border = borde
    for i, cap in enumerate([24, 48, 96]):
        r = 14 + i
        n = sum(1 for e in pend if e.capacidad == cap)
        for j, v in enumerate([f"{cap} hilos", n, n * cap]):
            c = rs.cell(r, 4 + j, v)
            c.border = borde
            if j:
                c.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFG", [36, 12, 3, 18, 13, 12, 12]):
        rs.column_dimensions[col].width = w

    # ---------- HOJA PENDIENTES ----------
    wp = wb.create_sheet("PENDIENTES")
    hdr_row(wp, 1, ["N°", "ANILLO", "ENLACE", "TRAMO (A → B)", "CAP.",
                    "CREADO", "DÍAS", "QUÉ FALTA"])
    for i, e in enumerate(sorted(pend, key=lambda x: x.creado), 1):
        wp.append([i, e.anillo, e.nombre,
                   f"{e.origen_a or '?'} → {e.origen_b or '?'}",
                   e.capacidad, e.creado.strftime("%d/%m/%Y %H:%M"),
                   dias(e), e.pendiente_texto])
        for c in range(1, 9):
            cell = wp.cell(i + 1, c)
            cell.border = borde
            cell.font = Font(size=9)
            if i % 2 == 0:
                cell.fill = gris
    for col, w in zip("ABCDEFGH", [5, 15, 30, 30, 7, 15, 7, 24]):
        wp.column_dimensions[col].width = w
    wp.freeze_panes = "A2"
    wp.auto_filter.ref = f"A1:H{len(pend)+1}"

    # ---------- HOJA VALIDADOS ----------
    wv = wb.create_sheet("VALIDADOS")
    hdr_row(wv, 1, ["N°", "ANILLO", "ENLACE", "TRAMO (A ⇄ B)", "CAP.",
                    "CREADO", "HILOS CON DIFERENCIAS"], verde)
    for i, e in enumerate(sorted(valid, key=lambda x: x.creado), 1):
        difs = sum(1 for h in e.hilos if not h.coincide)
        wv.append([i, e.anillo, e.nombre,
                   f"{e.origen_a} ⇄ {e.origen_b}", e.capacidad,
                   e.creado.strftime("%d/%m/%Y %H:%M"), difs])
        for c in range(1, 8):
            cell = wv.cell(i + 1, c)
            cell.border = borde
            cell.font = Font(size=9)
    for col, w in zip("ABCDEFG", [5, 15, 30, 30, 7, 15, 22]):
        wv.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"REPORTE_AVANCE_{ahora:%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


if __name__ == "__main__":

    app.run(debug=True)
